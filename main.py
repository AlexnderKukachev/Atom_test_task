# from bs4 import BeautifulSoup
# from requests import get, exceptions
#
#
# def count_tags():
#     number_of_tags = 0
#     tags_with_attrs = 0
#
#     url = 'https://jetlend.ru/'
#     try:
#         response = get(url)
#     except exceptions.ConnectionError:
#         print(f'Service - {url} is not available')
#         return
#     content = response.text
#     soup = BeautifulSoup(content, "html.parser")
#     for child in soup.recursiveChildGenerator():
#         if child.name:
#             number_of_tags += 1
#             if child.attrs:
#                 tags_with_attrs += 1
#     print(f'Tags on the page - {number_of_tags}')
#     print(f'Tags with attributes on the page - {tags_with_attrs}')
#
#
# if __name__ == '__main__':
#     count_tags()
#
#
# def get_ip() -> str:
#     from requests import get, exceptions
#     url = 'https://ifconfig.me/ip'
#     try:
#         response = get(url)
#     except exceptions.ConnectionError:
#         return f'Service - {url} is not available'
#     return response.content.decode()
#
#
# print(get_ip())


from openpyxl import load_workbook
from os import path, curdir
from datetime import datetime, timedelta


class DocumentDoesNotExist(Exception):
    pass


def give_path() -> str:
    doc_path = path.join(curdir, 'borrowers.xlsx')
    if not path.exists(doc_path):
        raise DocumentDoesNotExist(f'File "borrowers.xlsx" does not exists or not available')
    return doc_path


def open_doc(doc_path: str):
    wb = load_workbook(doc_path)
    ws = wb.active
    return ws


def print_amount_2021(ws):
    end = datetime(year=2021, day=31, month=12) + timedelta(days=1)
    row = 1
    total_amount = 0
    while ws.cell(row + 1, 2).value:
        row += 1
        d = datetime.strptime(ws.cell(row, 2).value, '%Y-%m-%d %H:%M:%S.%f')
        if d < end:
            total_amount += ws.cell(row, 4).value
    print(total_amount)


def print_rating_amount(ws):
    row = 1
    rating_amount = {}
    while ws.cell(row + 1, 3).value:
        row += 1
        value = ws.cell(row, 3).value
        if value not in rating_amount:
            rating_amount[value] = ws.cell(row, 4).value
        else:
            rating_amount[value] += ws.cell(row, 4).value
    print('rating\t\tamount')
    for key in rating_amount:
        key_p = str(key)
        if len(key_p) < len('rating'):
            for i in range(len('rating')-len(key_p)):
                key_p += ' '
        print(key_p, '\t\t', rating_amount[key])


if __name__ == '__main__':
    doc_path = give_path()
    ws = open_doc(doc_path)
    print_amount_2021(ws)
    print_rating_amount(ws)
